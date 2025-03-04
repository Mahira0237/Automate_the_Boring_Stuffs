import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

filename='TextFilesToSpreadsheet.xlsx'
wb = openpyxl.load_workbook(filename)
sheet=wb['Expenditure']

for i in range(1,sheet.max_row+1):
    f = open("I:\\My Drive\\AutomateTheBoringStuffs\\Chapter13\\SpreadsheetToTextFiles\\items.txt", "a")
    f.write(sheet['A'+str(i)].value)
f.close()
for i in range(1,sheet.max_row+1):
    f = open("I:\\My Drive\\AutomateTheBoringStuffs\\Chapter13\\SpreadsheetToTextFiles\\quantity.txt", "a")
    f.write(sheet['B'+str(i)].value)
f.close()
for i in range(1,sheet.max_row+1):
    f = open("I:\\My Drive\\AutomateTheBoringStuffs\\Chapter13\\SpreadsheetToTextFiles\\costs.txt", "a")
    f.write(sheet['C'+str(i)].value)
f.close()
for i in range(1,sheet.max_row+1):
    f = open("I:\\My Drive\\AutomateTheBoringStuffs\\Chapter13\\SpreadsheetToTextFiles\\total.txt", "a")
    f.write(sheet['D'+str(i)].value)
f.close()

