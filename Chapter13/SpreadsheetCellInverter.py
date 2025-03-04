import openpyxl
from openpyxl.utils import get_column_letter
filename='example.xlsx'
def InvertSheet(filename=str):
    filename2='example_inverted.xlsx'
    wb = openpyxl.load_workbook(filename)
    wb2 = openpyxl.Workbook()
    sheet=wb['Sheet1']
    sheet2=wb2.active
    for i in range(1,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            a = sheet.cell(row = i, column = j).value 
            sheet2.cell(row = j, column = i).value = a
    wb2.save(filename2)
InvertSheet(filename)