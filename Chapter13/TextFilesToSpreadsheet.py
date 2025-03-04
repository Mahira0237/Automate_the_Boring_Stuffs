import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
Bold=Font(bold=True)
Items=open('I:\\My Drive\\AutomateTheBoringStuffs\\Chapter13\\TextFilesToSpreadsheet\\fileItems.txt',"r").readlines()
Quantity=open('I:\\My Drive\\AutomateTheBoringStuffs\\Chapter13\\TextFilesToSpreadsheet\\fileQuanity.txt',"r").readlines()
Costs=open('I:\\My Drive\\AutomateTheBoringStuffs\\Chapter13\\TextFilesToSpreadsheet\\fileCost.txt',"r").readlines()
Total=open('I:\\My Drive\\AutomateTheBoringStuffs\\Chapter13\\TextFilesToSpreadsheet\\fileTotal.txt',"r").readlines()
NoRows=len(Items)
sheet=wb['Sheet']
sheet.title='Expenditure'
for i in range(1,NoRows+1):
    sheet[get_column_letter(i)+'1'].font = Bold
for i in range(1,NoRows+1):
    sheet.cell(row = i, column = 1).value = Items[i-1]
for i in range(1,NoRows+1):
    sheet.cell(row = i, column = 2).value = Quantity[i-1]
for i in range(1,NoRows+1):
    sheet.cell(row = i, column = 3).value = Costs[i-1]
for i in range(1,NoRows+1):
    sheet.cell(row = i, column = 4).value = Total[i-1]
wb.save('TextFilesToSpreadsheet.xlsx')
