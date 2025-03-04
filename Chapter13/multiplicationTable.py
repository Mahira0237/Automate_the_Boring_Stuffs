import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
wb = openpyxl.Workbook()
N=6
i=1
sheet=wb['Sheet']
sheet.title='Table'
Bold=Font(bold=True)
for rowNum in range(2,N+2):
    sheet['A'+str(rowNum)].font = Bold
    sheet['A'+str(rowNum)]=rowNum-1    
    sheet[get_column_letter(rowNum)+'1'].font=Bold
    sheet[get_column_letter(rowNum)+'1']=rowNum-1 
for rows in sheet['B'+str(2):get_column_letter(N+1)+str(N+1)]:
    for objects in rows:
        objects.value=((objects.column)-1)*i
    i+=1

wb.save('multiplcationTable.xlsx')