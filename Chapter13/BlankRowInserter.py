import openpyxl
filename='example.xlsx'
N=2
M=3
def BlankRow(N=int,M=int,filename=str):
    filename2='example_withblanks.xlsx'
    wb = openpyxl.load_workbook(filename)
    wb2 = openpyxl.Workbook()
    sheet=wb['Sheet1']
    sheet2=wb2.active
    start=N+1
    for i in range(1,start):
        for j in range(1,sheet.max_column+1):
            a = sheet.cell(row = i, column = j).value 
            sheet2.cell(row = i, column = j).value = a
    for i in range(start,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            c = sheet.cell(row = i, column = j).value 
            sheet2.cell(row = i+M, column = j).value = c
    wb2.save(filename2)
BlankRow(N,M,filename)
